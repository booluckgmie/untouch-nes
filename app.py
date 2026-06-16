# ─────────────────────────────────────────────────────────────────────────────
# app.py  — NES Analytics Dashboard  (Flask backend)
# Run:  python app.py  |  Open: http://localhost:5001
# ─────────────────────────────────────────────────────────────────────────────
import json
import logging
import threading
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from flask import Flask, jsonify, render_template, request

import cache as C
import nes_db as DB

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s %(levelname)s %(name)s: %(message)s")
_log = logging.getLogger(__name__)

app = Flask(__name__)

# ── Active subcategories ──────────────────────────────────────────────────────
# sub01 NADI4U-Entrepreneur | sub02 NADI4U-LifelongLearning
# sub03 NADI4U-Wellbeing    | sub08 NADI2U-Wellbeing
# sub04-07, sub09-10 are not part of this reporting scope.
ACTIVE_SUBS = {1, 2, 3, 4, 8}

# ── Static 1,099 NADI index — loaded from data/all_sites.json ────────────────
_SITES_FILE   = Path(__file__).parent / "data" / "all_sites.json"
_STATIC_SITES = json.loads(_SITES_FILE.read_text(encoding="utf-8"))["sites"]
assert len(_STATIC_SITES) == 1099, f"Expected 1099 sites, got {len(_STATIC_SITES)}"

# ── In-memory progress tracker ────────────────────────────────────────────────
_progress: dict = {}  # str(sub_id) → {status, pct, msg, mod, chunks, done_chunks}


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
def _set_prog(sid, pct, msg, **kw):
    prev = _progress.get(sid, {})
    _progress[sid] = {"status": "running", "pct": pct, "msg": msg,
                      "mod": prev.get("mod", ""), **kw}


def _get_sites() -> list:
    """Return 1,099-row NADI index — no DB needed."""
    return _STATIC_SITES


def _fetch_and_cache(sub_id: int, start_date: str | None = None, end_date: str | None = None,
                     programs: list | None = None):
    """Background thread: fetch one subcategory from DB and cache all blobs."""
    sid  = str(sub_id)
    info = DB.SUBCATEGORIES.get(sub_id, {})
    date_label = ""
    if start_date or end_date:
        date_label = f" [{start_date or '…'} → {end_date or '…'}]"
    _progress[sid] = {"status": "running", "pct": 3, "msg": f"Connecting to DB{date_label} …",
                      "mod": info.get("mod", ""), "chunks": 0, "done_chunks": 0}
    try:
        _scols = ["refid_mcmc", "nadi_name", "state", "tp", "dusp"]
        df_sites = pd.DataFrame(_STATIC_SITES, columns=_scols[:len(_STATIC_SITES[0]) if _STATIC_SITES else 3])

        _set_prog(sid, 7, "Getting event IDs …")
        from db_configs import get_conn
        conn = get_conn()
        try:
            event_ids = DB._get_event_ids(conn, sub_id, start_date, end_date)
        finally:
            conn.close()

        if not event_ids:
            _progress[sid] = {"status": "empty", "pct": 100,
                               "msg": "No events for this subcategory.",
                               "mod": info.get("mod", "")}
            return

        chunks   = [event_ids[i:i + DB._EVENT_CHUNK]
                    for i in range(0, len(event_ids), DB._EVENT_CHUNK)]
        n_chunks = len(chunks)
        _set_prog(sid, 10,
                  f"{len(event_ids):,} events → {n_chunks} chunks",
                  chunks=n_chunks, done_chunks=0)

        ev_frames, par_frames = [], []
        conn = get_conn()
        try:
            for idx, chunk in enumerate(chunks, 1):
                pct = 10 + int(75 * idx / n_chunks)
                _set_prog(sid, pct,
                          f"Chunk {idx}/{n_chunks} · {idx * DB._EVENT_CHUNK:,} events",
                          chunks=n_chunks, done_chunks=idx)
                ev_frames.append(DB._fetch_events(conn, chunk))
                par_frames.append(DB._fetch_participants(conn, chunk,
                                                         f"nes_{sub_id}_{idx}"))
                conn.commit()
        finally:
            conn.close()

        _set_prog(sid, 86, "Merging data …")
        df_ev  = pd.concat(ev_frames,  ignore_index=True).drop_duplicates("event_id")
        df_par = pd.concat(par_frames, ignore_index=True)
        df     = df_par.merge(df_ev, on="event_id", how="left")
        _log.info("sub%d: %d rows | %d events", sub_id, len(df), df["event_id"].nunique())

        _set_prog(sid, 88, "Applying SSO remap …")
        df = DB.apply_sso_remap(df, sub_id)

        if programs:
            before = len(df)
            df = df[df["program"].isin(programs)].reset_index(drop=True)
            _log.info("Program filter %s: %d → %d rows", programs, before, len(df))

        _set_prog(sid, 90, "Building NADI index …")
        nadi_idx = DB.build_nadi_index(df, df_sites, sub_id)

        _set_prog(sid, 94, "Building monthly data …")
        monthly  = DB.build_monthly(df, df_sites, sub_id)

        _set_prog(sid, 97, "Building weekly data …")
        weekly   = DB.build_weekly(df, df_sites, sub_id)

        _set_prog(sid, 99, "Saving …")
        C.set(f"sub_{sub_id}", {
            "sub_id":          sub_id,
            "fetched":         datetime.now().strftime("%d %b %Y, %H:%M"),
            "start_date":      start_date,
            "end_date":        end_date,
            "programs_filter": programs,
            "nadi":            nadi_idx,
            "monthly":         monthly,
            "weekly":          weekly,
        })
        _progress[sid] = {"status": "done", "pct": 100, "msg": "Ready ✓",
                          "mod": info.get("mod", ""),
                          "chunks": n_chunks, "done_chunks": n_chunks}
        _log.info("sub%d cached ✓", sub_id)

    except Exception as exc:
        import traceback; traceback.print_exc()
        _log.error("sub%d fetch failed: %s", sub_id, exc)
        _progress[sid] = {"status": "error", "pct": 0, "msg": str(exc),
                          "mod": info.get("mod", "")}


# ─────────────────────────────────────────────────────────────────────────────
# Routes
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/ping")
def api_ping():
    cached = [f for f in Path("data").iterdir()
              if f.suffix == ".json" and not f.name.endswith("meta.json")]
    return jsonify({"ok": True, "status": "running",
                    "active_subs": sorted(ACTIVE_SUBS),
                    "cached": len(cached)})


@app.route("/api/sites")
def api_sites():
    return jsonify({"ok": True, "sites": _STATIC_SITES})


@app.route("/api/subcategories")
def api_subcategories():
    result = []
    for sub_id in sorted(ACTIVE_SUBS):
        info   = DB.SUBCATEGORIES.get(sub_id, {})
        sid    = str(sub_id)
        cached = C.get(f"sub_{sub_id}")
        result.append({
            "id":       sub_id,
            "cat":      info.get("cat", ""),
            "mod":      info.get("mod", ""),
            "label":    info.get("label", ""),
            "cached":   cached is not None,
            "fetched":  cached["fetched"] if cached else None,
            "progress": _progress.get(sid, {}),
        })
    return jsonify({"ok": True, "subcategories": result})


@app.route("/api/fetch/<int:sub_id>", methods=["POST"])
def api_fetch(sub_id: int):
    if sub_id not in ACTIVE_SUBS:
        return jsonify({"ok": False, "error": "Not an active subcategory"}), 400
    sid = str(sub_id)
    if _progress.get(sid, {}).get("status") == "running":
        return jsonify({"ok": False, "error": "Already fetching"}), 409
    body       = request.get_json(silent=True) or {}
    start_date = body.get("start_date") or None
    end_date   = body.get("end_date")   or None
    programs   = body.get("programs")   or None
    threading.Thread(target=_fetch_and_cache,
                     args=(sub_id, start_date, end_date, programs), daemon=True).start()
    return jsonify({"ok": True, "msg": f"Fetch started for sub{sub_id}"})


@app.route("/api/progress/<int:sub_id>")
def api_progress(sub_id: int):
    return jsonify(_progress.get(str(sub_id),
                                 {"status": "idle", "pct": 0, "msg": "Not started"}))


@app.route("/api/data/<int:sub_id>")
def api_data(sub_id: int):
    cached = C.get(f"sub_{sub_id}")
    if not cached:
        return jsonify({"ok": False, "error": "Not cached — fetching in background"}), 404
    return jsonify({"ok": True, "data": cached})


@app.route("/api/cache/status")
def api_cache_status():
    return jsonify({"ok": True, "cache": C.list_cached()})


@app.route("/api/cache/invalidate/<int:sub_id>", methods=["POST"])
def api_cache_invalidate(sub_id: int):
    C.invalidate(f"sub_{sub_id}")
    _progress.pop(str(sub_id), None)
    return jsonify({"ok": True})


@app.route("/api/cache/invalidate_all", methods=["POST"])
def api_cache_invalidate_all():
    C.invalidate_all()
    _progress.clear()
    return jsonify({"ok": True})


@app.route("/api/template_xlsx")
def api_template_xlsx():
    from flask import send_file
    tmpl = Path(__file__).parent / "template_report.xlsx"
    return send_file(str(tmpl),
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/api/export/<int:sub_id>")
def api_export(sub_id: int):
    """Generate Excel report in MCMC/TM Data Smart Services template format."""
    import io
    import openpyxl
    from flask import send_file

    cached = C.get(f"sub_{sub_id}")
    if not cached:
        return jsonify({"ok": False, "error": "Data not cached — fetch first"}), 404

    monthly    = cached.get("monthly", {})
    months_lst = monthly.get("months", [])   # ["Jan 2026", …]
    rows       = monthly.get("rows", [])     # [[nadi_idx,[ev…],[pax…],{pk:{ev,px}}],…]

    # Map each data-list index → 0-based month offset (0=Jan … 11=Dec)
    MN = {'Jan':0,'Feb':1,'Mar':2,'Apr':3,'May':4,'Jun':5,
          'Jul':6,'Aug':7,'Sep':8,'Oct':9,'Nov':10,'Dec':11}
    month_map = {}
    for i, m in enumerate(months_lst):
        mn = (m or '').split()[0][:3]
        if mn in MN:
            month_map[i] = MN[mn]

    # 16 states in template row order (rows 6–21)
    STATES = ['JOHOR','KEDAH','KELANTAN','MELAKA','NEGERI SEMBILAN',
              'PAHANG','PERAK','PERLIS','PULAU PINANG','SABAH',
              'SARAWAK','SELANGOR','TERENGGANU','W.P KUALA LUMPUR',
              'W.P LABUAN','W.P PUTRAJAYA']

    # Aggregate total events per state × month (0-indexed)
    state_ev = {s: [0]*12 for s in STATES}
    for row in rows:
        if not row or len(row) < 2:
            continue
        nadi_idx = row[0]
        tot_ev   = row[1] if isinstance(row[1], list) else []
        if nadi_idx < len(_STATIC_SITES):
            state = str(_STATIC_SITES[nadi_idx][2]).upper().strip()
            if state in state_ev:
                for data_i, col_off in month_map.items():
                    if data_i < len(tot_ev):
                        state_ev[state][col_off] += int(tot_ev[data_i] or 0)

    # Load template — preserves all cell styles, merges, colours
    tmpl = Path(__file__).parent / "template_report.xlsx"
    wb   = openpyxl.load_workbook(str(tmpl))
    ws   = wb.active

    # Fill ALL-NADI events: cols C–N (openpyxl col 3–14), rows 6–21
    for si, state in enumerate(STATES):
        row_num = 6 + si
        ev = state_ev[state]
        for mo in range(12):
            ws.cell(row=row_num, column=3 + mo).value = ev[mo] if ev[mo] else None
        ws.cell(row=row_num, column=15).value = sum(ev) or None   # col O = TOTAL

    # Row 22 — TOTAL BY MONTH
    for mo in range(12):
        total = sum(state_ev[s][mo] for s in STATES)
        ws.cell(row=22, column=3 + mo).value = total if total else None
    grand = sum(sum(state_ev[s]) for s in STATES)
    ws.cell(row=22, column=15).value = grand or None

    # Row 23 — TOTAL BY PROGRAMME (merged C23:N23 — write to first cell)
    ws.cell(row=23, column=3).value = grand or None

    info  = DB.SUBCATEGORIES.get(sub_id, {})
    mod   = info.get("mod", f"Sub{sub_id}")
    start = cached.get("start_date") or "ALL"
    end   = cached.get("end_date")   or str(date.today())
    fname = f"NES_{mod}_{start}_{end}.xlsx"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name=fname)


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("\n  NES Analytics Dashboard  ->  http://localhost:5001\n", flush=True)
    app.run(debug=True, host="0.0.0.0", port=5001, threaded=True)
